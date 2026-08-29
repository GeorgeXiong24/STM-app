from __future__ import annotations

import shutil
import sys
import tempfile
import random
import json
import html
import re
import os
import subprocess
import ssl
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any


REQUIRED_PACKAGES = (
    "PySide6>=6.7",
    "numbers-parser>=4.19",
    "openpyxl>=3.1",
)
REQUIRED_MODULES = ("PySide6", "numbers_parser", "openpyxl")


def _ensure_runtime_environment() -> None:
    if getattr(sys, "frozen", False):
        return

    project_directory = Path(__file__).resolve().parent
    environment_directory = project_directory / ".venv"
    in_virtual_environment = sys.prefix != sys.base_prefix
    environment_python = (
        Path(sys.executable)
        if in_virtual_environment
        else environment_directory / "bin" / "python"
    )

    try:
        if not environment_python.exists():
            subprocess.check_call(
                [sys.executable, "-m", "venv", str(environment_directory)]
            )

        missing_modules = subprocess.run(
            [
                str(environment_python),
                "-c",
                "import " + ", ".join(REQUIRED_MODULES),
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        ).returncode != 0
        if not missing_modules:
            if in_virtual_environment:
                return
            os.execv(
                str(environment_python),
                [str(environment_python), str(Path(__file__).resolve()), *sys.argv[1:]],
            )

        subprocess.check_call(
            [
                str(environment_python),
                "-m",
                "pip",
                "install",
                "--disable-pip-version-check",
                *REQUIRED_PACKAGES,
            ]
        )
    except (OSError, subprocess.CalledProcessError) as error:
        raise SystemExit(
            "STM could not prepare its Python environment. "
            "Please check that Python 3.10+ and an internet connection are available.\n"
            f"Details: {error}"
        ) from error

    if not in_virtual_environment:
        os.execv(
            str(environment_python),
            [str(environment_python), str(Path(__file__).resolve()), *sys.argv[1:]],
        )


_ensure_runtime_environment()

from numbers_parser import Document
from numbers_parser import Document as NumbersDocument
from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QObject, QThread, QTimer, Qt, Signal, Slot

key = ""
from PySide6.QtGui import QAction
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QHeaderView,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QProgressDialog,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)


class AnswerJudgeWorker(QObject):
    finished = Signal(object, str)

    def __init__(self, word: str, explanation: str, answer: str) -> None:
        super().__init__()
        self.word = word
        self.explanation = explanation
        self.answer = answer

    @staticmethod
    def _build_ssl_context(base_url: str) -> ssl.SSLContext:
        verify_ssl = os.environ.get("DEEPSEEK_VERIFY_SSL", "true").lower()
        should_verify = verify_ssl not in {"0", "false", "no", "off"}

        if should_verify:
            return ssl.create_default_context()

        return ssl._create_unverified_context()

    @staticmethod
    def _is_certificate_error(error: BaseException) -> bool:
        message = str(error).lower()
        return (
            "certificate verify failed" in message
            or "self-signed certificate" in message
            or "ssl: cert" in message
            or "unable to get local issuer certificate" in message
        )

    def _send_judgment_request(self, request: urllib.request.Request, ssl_context: ssl.SSLContext):
        with urllib.request.urlopen(request, timeout=30, context=ssl_context) as response:
            result = json.loads(response.read().decode("utf-8"))
        content = result["choices"][0]["message"]["content"].strip()
        if content.startswith("```"):
            content = content.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        judgment = json.loads(content)
        correct = judgment["correct"]
        reason = str(judgment.get("reason", ""))
        if not isinstance(correct, bool):
            raise ValueError("The API returned a non-boolean correctness value.")
        return correct, reason

    def run(self) -> None:
        api_key = (key or "").strip()
        if not api_key:
            self.finished.emit(False, "DEEPSEEK_API_KEY is not set.")
            return

        base_url = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
        model = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
        payload = {
            "model": model,
            "temperature": 0,
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "Judge whether a student's Chinese definition matches the reference "
                        "meaning. Accept synonyms, natural paraphrases, and minor typos. "
                        "Reject unrelated or materially incorrect meanings. Return JSON only "
                        "with boolean 'correct' and short Chinese string 'reason'."
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "word": self.word,
                            "reference_definition": self.explanation,
                            "student_answer": self.answer,
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        }
        request = urllib.request.Request(
            f"{base_url.rstrip('/')}/chat/completions",
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        ssl_context = self._build_ssl_context(base_url)
        try:
            correct, reason = self._send_judgment_request(request, ssl_context)
        except urllib.error.HTTPError as error:
            try:
                details = error.read().decode("utf-8", errors="replace")
            except Exception:
                details = str(error)
            self.finished.emit(None, f"DeepSeek HTTP {error.code}: {details}")
            return
        except urllib.error.URLError as error:
            if self._is_certificate_error(error):
                try:
                    correct, reason = self._send_judgment_request(request, ssl._create_unverified_context())
                except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, KeyError, IndexError, ValueError, ssl.SSLError) as retry_error:
                    self.finished.emit(None, f"Could not judge the answer: {retry_error}")
                    return
                self.finished.emit(correct, reason)
                return
            self.finished.emit(None, f"Could not judge the answer: {error}")
            return
        except (TimeoutError, KeyError, IndexError, ValueError, ssl.SSLError) as error:
            if self._is_certificate_error(error):
                try:
                    correct, reason = self._send_judgment_request(request, ssl._create_unverified_context())
                except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, KeyError, IndexError, ValueError, ssl.SSLError) as retry_error:
                    self.finished.emit(None, f"Could not judge the answer: {retry_error}")
                    return
                self.finished.emit(correct, reason)
                return
            self.finished.emit(None, f"Could not judge the answer: {error}")
            return
        self.finished.emit(correct, reason)


class SentenceWorker(QObject):
    finished = Signal(str, str)

    def __init__(self, word: str, explanation: str, refresh: bool = False) -> None:
        super().__init__()
        self.word = word
        self.explanation = explanation
        self.refresh = refresh

    def run(self) -> None:
        api_key = (key or "").strip()
        if not api_key:
            self.finished.emit("", "DEEPSEEK_API_KEY is not set.")
            return

        base_url = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
        model = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
        instruction = (
            "Create one natural, concise English example sentence using the given word and "
            "matching its Chinese meaning. You may change the word's tense, number, or form "
            "when natural. Wrap the complete word form used in the sentence in double brackets "
            "like [[word]]. Return only the sentence."
        )
        if self.refresh:
            instruction = (
                "Create a new and different natural English example sentence using the given "
                "word and matching its Chinese meaning. You may change the word's tense, "
                "number, or form when natural. Use a different context and wording from any "
                "previous example. Wrap the complete word form used in the sentence in double "
                "brackets like [[word]]. Return only the sentence."
            )
        payload = {
            "model": model,
            "temperature": 0.8 if self.refresh else 0.3,
            "messages": [
                {
                    "role": "system",
                    "content": instruction,
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {"word": self.word, "chinese_meaning": self.explanation},
                        ensure_ascii=False,
                    ),
                },
            ],
        }
        request = urllib.request.Request(
            f"{base_url.rstrip('/')}/chat/completions",
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(
                request,
                timeout=30,
                context=AnswerJudgeWorker._build_ssl_context(base_url),
            ) as response:
                result = json.loads(response.read().decode("utf-8"))
            sentence = result["choices"][0]["message"]["content"].strip()
            if sentence.startswith("```"):
                sentence = sentence.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            if not sentence:
                raise ValueError("DeepSeek returned an empty sentence.")
        except urllib.error.HTTPError as error:
            try:
                details = error.read().decode("utf-8", errors="replace")
            except Exception:
                details = str(error)
            self.finished.emit("", f"DeepSeek HTTP {error.code}: {details}")
            return
        except urllib.error.URLError as error:
            if AnswerJudgeWorker._is_certificate_error(error):
                try:
                    with urllib.request.urlopen(
                        request,
                        timeout=30,
                        context=ssl._create_unverified_context(),
                    ) as response:
                        result = json.loads(response.read().decode("utf-8"))
                    sentence = result["choices"][0]["message"]["content"].strip()
                except (urllib.error.HTTPError, urllib.error.URLError, KeyError, IndexError, TypeError, ValueError, json.JSONDecodeError) as retry_error:
                    self.finished.emit("", f"Could not generate a sentence: {retry_error}")
                    return
                self.finished.emit(sentence, "")
                return
            self.finished.emit("", f"Could not generate a sentence: {error}")
            return
        except (KeyError, IndexError, TypeError, ValueError, json.JSONDecodeError, ssl.SSLError) as error:
            self.finished.emit("", f"Could not generate a sentence: {error}")
            return
        self.finished.emit(sentence, "")


class SpreadsheetWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.current_path: Path | None = None
        self.temporary_directory = tempfile.TemporaryDirectory(prefix="STM-")
        self.setWindowTitle("STM")
        self.setMinimumSize(900, 760)
        available_geometry = QApplication.primaryScreen().availableGeometry()
        initial_width = max(900, int(available_geometry.width() * 0.8))
        initial_height = min(980, int(available_geometry.height() * 0.8))
        self.resize(initial_width, max(760, initial_height))
        self.key = ""
        self._build_ui()
        self._apply_styles()
        self._show_key_entry_screen()

    def _build_ui(self) -> None:
        self.content_widget = QWidget()
        root = QVBoxLayout(self.content_widget)
        root.setContentsMargins(34, 28, 34, 28)
        root.setSpacing(20)

        header = QHBoxLayout()
        brand = QLabel("STM")
        brand.setObjectName("brand")
        header.addWidget(brand)
        header.addStretch()
        self.open_button = QPushButton("Open sheet")
        self.open_button.setObjectName("primaryButton")
        self.open_button.clicked.connect(self.choose_file)
        header.addWidget(self.open_button)
        root.addLayout(header)

        intro = QVBoxLayout()
        intro.setSpacing(5)
        title = QLabel("Your workspace for sheets.")
        title.setObjectName("title")
        subtitle = QLabel("Upload a Numbers or Excel file of your word list.")
        subtitle.setObjectName("subtitle")
        intro.addWidget(title)
        intro.addWidget(subtitle)
        root.addLayout(intro)

        self.drop_zone = QFrame()
        self.drop_zone.setObjectName("dropZone")
        self.drop_zone.setAcceptDrops(True)
        drop_layout = QVBoxLayout(self.drop_zone)
        drop_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_layout.setSpacing(10)
        upload_mark = QLabel("↑")
        upload_mark.setObjectName("uploadMark")
        upload_mark.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_layout.addWidget(upload_mark)
        drop_title = QLabel("Drop a sheet here")
        drop_title.setObjectName("dropTitle")
        drop_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_layout.addWidget(drop_title)
        drop_hint = QLabel(".numbers and .xlsx files")
        drop_hint.setObjectName("dropHint")
        drop_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_layout.addWidget(drop_hint)
        self.drop_zone.dragEnterEvent = self._drag_enter
        self.drop_zone.dropEvent = self._drop_file
        root.addWidget(self.drop_zone)

        self.file_label = QLabel("No file uploaded")
        self.file_label.setObjectName("fileLabel")
        root.addWidget(self.file_label)

        self.words_label = QLabel("Words")
        self.words_label.setObjectName("wordsLabel")
        root.addWidget(self.words_label)
        self.words_table = QTableWidget(0, 3)
        self.words_table.setObjectName("wordsTable")
        self.words_table.setHorizontalHeaderLabels(["Words", "Definitions", ""])
        self.words_table.setAlternatingRowColors(True)
        self.words_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.words_table.horizontalHeader().setStretchLastSection(False)
        self.words_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.words_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.words_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        root.addWidget(self.words_table, 1)

        self.storage_label = QLabel("Uploaded files are removed when you close STM.")
        self.storage_label.setObjectName("storageLabel")
        root.addWidget(self.storage_label)

        self.status_label = QLabel("Waiting for a file to be uploaded...")
        self.status_label.setObjectName("statusLabel")
        root.addWidget(self.status_label)

        bottom_actions = QHBoxLayout()
        self.back_button = QPushButton("Back")
        self.back_button.setObjectName("goButton")
        self.back_button.clicked.connect(self._return_to_key_entry)
        bottom_actions.addWidget(self.back_button)
        bottom_actions.addStretch()
        self.go_button = QPushButton("Go")
        self.go_button.setObjectName("goButton")
        self.go_button.setProperty("hasFile", False)
        self.go_button.clicked.connect(self._clear_window)
        bottom_actions.addWidget(self.go_button)
        root.addLayout(bottom_actions)

        file_menu = self.menuBar().addMenu("File")
        open_action = QAction("Open…", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self.choose_file)
        file_menu.addAction(open_action)

    def _show_key_entry_screen(self) -> None:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(180, 220, 180, 200)
        layout.setSpacing(18)

        title = QLabel("Hi! This is STM by George Xiong.")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        row = QHBoxLayout()
        self.key_input = QLineEdit()
        self.key_input.setPlaceholderText("Please input the key")
        self.key_input.setObjectName("answerBox")
        self.key_input.setMinimumWidth(620)
        self.key_input.setFixedHeight(60)
        self.key_input.returnPressed.connect(self._submit_key)
        row.addWidget(self.key_input)

        self.key_button = QPushButton("OK")
        self.key_button.setObjectName("primaryButton")
        self.key_button.clicked.connect(self._submit_key)
        row.addWidget(self.key_button)
        layout.addLayout(row)

        self.key_entry_widget = widget
        self._set_central_widget(widget)
        self.key_input.setFocus(Qt.FocusReason.OtherFocusReason)

    def _set_central_widget(self, widget: QWidget) -> None:
        current_widget = self.takeCentralWidget()
        if current_widget is not None and current_widget is not widget:
            current_widget.setParent(None)
        self.setCentralWidget(widget)

    def _show_standard_alert(self, title: str, message: str) -> None:
        box = QMessageBox(self)
        box.setWindowTitle(title)
        box.setText(message)
        box.setIcon(QMessageBox.Icon.NoIcon)
        box.setStandardButtons(QMessageBox.StandardButton.Ok)
        box.setStyleSheet("""
            QMessageBox {
                background: #f6f5f1;
                border: 1px solid #e1ddd4;
                border-radius: 12px;
                min-width: 420px;
            }
            QMessageBox QLabel {
                color: #20211f;
                font: 500 15px 'Avenir Next';
                margin: 0px;
            }
            QMessageBox QPushButton {
                background: #c85132;
                color: white;
                border: 0;
                border-radius: 6px;
                padding: 10px 22px;
                min-width: 90px;
                font: 600 13px 'Avenir Next';
            }
            QMessageBox QPushButton:hover {
                background: #ad4127;
            }
            QMessageBox > QPushButton {
                margin-top: 12px;
            }
            QMessageBox::text {
                padding-right: 0;
            }
            QMessageBox::button-layout {
                min-height: 42px;
            }
        """)
        box.setDefaultButton(QPushButton("OK"))
        box.show()
        QTimer.singleShot(1750, box.accept)

    def _submit_key(self) -> None:
        entered_key = self.key_input.text().strip()
        if not entered_key:
            self._show_standard_alert("Missing key", "Please input the key before continuing.")
            return
        if not entered_key.startswith("sk-"):
            self._show_standard_alert("Invalid key", "The input key is invalid, please redo.")
            self.key_input.clear()
            self.key_input.setFocus()
            return

        global key
        key = entered_key
        self.key = entered_key
        self._set_central_widget(self.content_widget)

    def _return_to_key_entry(self) -> None:
        previous_key = self.key or key
        if self.current_path is not None:
            try:
                self.current_path.unlink(missing_ok=True)
            except OSError:
                pass
        self.current_path = None
        self.words_table.setRowCount(0)
        self.words_label.setText("Words")
        self.file_label.setText("No file uploaded")
        self.storage_label.setText("Uploaded files are removed when you close STM.")
        self.status_label.setText("Waiting for a file to be uploaded...")
        self._set_go_file_state(False)
        self._show_key_entry_screen()
        self.key_input.setText(previous_key)
        self.key_input.selectAll()
        self.key_input.setFocus(Qt.FocusReason.OtherFocusReason)

    def _apply_styles(self) -> None:
        self.setStyleSheet("""
            QMainWindow, QWidget { background: #f6f5f1; color: #20211f; }
            QMenuBar { background: #f6f5f1; padding: 5px 12px; }
            QMenuBar::item:selected { background: #e8e5dc; }
            #brand { color: #c85132; font: 700 27px 'Avenir Next'; letter-spacing: 2px; }
            #title { color: #20211f; font: 700 30px 'Avenir Next'; }
            #subtitle { color: #70736d; font: 15px 'Avenir Next'; }
            #primaryButton { background: #c85132; color: white; border: 0; border-radius: 6px; padding: 11px 19px; font: 600 14px 'Avenir Next'; }
            #primaryButton:hover { background: #ad4127; }
            #goButton { background: #20211f; color: white; border: 0; border-radius: 6px; padding: 10px 22px; font: 600 14px 'Avenir Next'; }
            #goButton[hasFile="false"] { background: #aaa9a3; }
            #goButton[hasFile="false"]:hover { background: #989791; }
            #goButton[hasFile="true"] { background: #c85132; }
            #goButton[hasFile="true"]:hover { background: #ad4127; }
            #dropZone { min-height: 145px; border: 1.5px dashed #c9c1b5; border-radius: 8px; background: #fbfaf7; }
            #dropZone:hover { border-color: #c85132; background: #fffdf9; }
            #uploadMark { color: #c85132; font: 700 30px 'Avenir Next'; }
            #dropTitle { color: #353633; font: 600 16px 'Avenir Next'; }
            #dropHint, #statusLabel, #storageLabel { color: #85877f; font: 13px 'Avenir Next'; }
            #progressLabel { color: #85877f; font: 600 14px 'Avenir Next'; }
            #reportHint { color: #85877f; font: 12px 'Avenir Next'; padding-top: 4px; }
            #fileLabel { color: #3f413d; font: 600 14px 'Avenir Next'; padding: 2px 0; }
            #wordsLabel { color: #353633; font: 600 16px 'Avenir Next'; padding-top: 4px; }
            #wordsTable { border: 1px solid #e1ddd4; border-radius: 6px; background: #fffdf9; font: 14px 'Avenir Next'; }
            #practiceWord { border: 1px solid #e1ddd4; border-radius: 8px; background: #fffdf9; font: 700 34px 'Avenir Next'; }
            #answerBox { border: 1px solid #c9c1b5; border-radius: 6px; background: #fffdf9; padding: 10px; font: 700 20px 'Avenir Next'; }
        """)

    def choose_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Open spreadsheet", "", "Sheets (*.numbers *.xlsx)")
        if path:
            self.open_file(Path(path))

    def _drag_enter(self, event: Any) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def _drop_file(self, event: Any) -> None:
        urls = event.mimeData().urls()
        if urls:
            self.open_file(Path(urls[0].toLocalFile()))
            event.acceptProposedAction()

    def _show_analysis_popup(self) -> None:
        self.analysis_popup = QProgressDialog(
            "Uploading and analyzing the file...", "", 0, 0, self
        )
        cancel_button = self.analysis_popup.findChild(QPushButton)
        if cancel_button is not None:
            cancel_button.hide()
        self.analysis_popup.setWindowTitle("Please wait")
        self.analysis_popup.setWindowModality(Qt.WindowModality.WindowModal)
        self.analysis_popup.setAutoClose(False)
        self.analysis_popup.setMinimumWidth(360)
        self.analysis_popup.setStyleSheet("""
            QProgressDialog {
                background: #f6f5f1;
                border: 1px solid #e1ddd4;
                border-radius: 12px;
            }
            QProgressDialog QLabel {
                color: #20211f;
                font: 500 15px 'Avenir Next';
                padding: 12px 18px;
            }
            QProgressBar {
                min-height: 8px;
                max-height: 8px;
                border: 0;
                border-radius: 4px;
                background: #e1ddd4;
            }
            QProgressBar::chunk {
                border-radius: 4px;
                background: #c85132;
            }
        """)
        self.analysis_popup.show()
        QApplication.processEvents()

    def _close_analysis_popup(self) -> None:
        popup = getattr(self, "analysis_popup", None)
        if popup is not None:
            popup.close()
            popup.deleteLater()
            self.analysis_popup = None

    def open_file(self, path: Path) -> None:
        if path.suffix.lower() not in {".numbers", ".xlsx"}:
            self._show_error("STM accepts .numbers and .xlsx files only.")
            return
        self._show_analysis_popup()
        try:
            temporary_path = Path(self.temporary_directory.name) / path.name
            shutil.copy2(path, temporary_path)
        except Exception as error:
            self._close_analysis_popup()
            self._show_error(f"Could not upload this file.\n\n{error}")
            return
        self.current_path = temporary_path
        self.file_label.setText(f"Uploaded: {path.name}")
        self.storage_label.setText(f"Temporarily stored as {temporary_path.name}")
        try:
            entries = self._read_words(temporary_path)
        except Exception as error:
            self.words_table.setRowCount(0)
            self.words_label.setText("Words")
            self._set_go_file_state(False)
            self._close_analysis_popup()
            self._show_error(f"Could not read the 单词 column.\n\n{error}")
            return

        self._close_analysis_popup()
        self.words_table.setRowCount(0)
        for word, explanation in entries:
            row = self.words_table.rowCount()
            self.words_table.insertRow(row)
            self.words_table.setItem(row, 0, QTableWidgetItem(word))
            self.words_table.setItem(row, 1, QTableWidgetItem(explanation))
            delete_button = QPushButton("delete")
            delete_button.clicked.connect(self._delete_word_row)
            self.words_table.setCellWidget(row, 2, delete_button)
        self.words_label.setText(f"Words ({len(entries)})")
        self._set_go_file_state(True)
        self.status_label.setText("Ready to start testing")

    def _clear_window(self) -> None:
        if not self.go_button.property("hasFile"):
            self._show_standard_alert("No file", "No file is uploaded.")
            return
        entries = [
            (word_item.text(), explanation_item.text())
            for row in range(self.words_table.rowCount())
            if (word_item := self.words_table.item(row, 0)) is not None
            and (explanation_item := self.words_table.item(row, 1)) is not None
        ]
        if not entries:
            self._show_standard_alert("No words", "There are no words.")
            return

        random.shuffle(entries)
        self.practice_entries = entries
        self.practice_incorrect_entries: list[tuple[str, str]] = []
        self.practice_given_up_entries: list[tuple[str, str]] = []
        self.practice_error_counts = {entry: 0 for entry in entries}
        self.practice_sentence_hints: dict[tuple[str, str], str] = {}
        self.practice_index = 0
        self.practice_attempts = 0
        self.practice_reviewing = False
        self._show_practice_window(self.practice_entries[0])
        self.menuBar().hide()

    def _show_practice_window(self, entry: tuple[str, str]) -> None:
        word, explanation = entry
        practice_widget = QWidget()
        practice_layout = QVBoxLayout(practice_widget)
        practice_layout.setContentsMargins(34, 28, 34, 28)
        practice_layout.setSpacing(20)

        progress_label = QLabel(
            f"{self.practice_index + 1}/{len(self.practice_entries)}"
        )
        progress_label.setObjectName("progressLabel")
        progress_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
        practice_layout.addWidget(progress_label)

        word_label = QLabel("Word")
        word_label.setObjectName("wordsLabel")
        word_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        practice_layout.addWidget(word_label)
        word_display = QLabel(word)
        word_display.setObjectName("practiceWord")
        word_display.setAlignment(Qt.AlignmentFlag.AlignCenter)
        word_display.setFixedSize(560, 150)
        practice_layout.addWidget(word_display, 0, Qt.AlignmentFlag.AlignHCenter)
        sentence_label = QLabel("")
        sentence_label.setObjectName("sentenceLabel")
        sentence_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sentence_label.setWordWrap(True)
        sentence_label.setTextFormat(Qt.TextFormat.RichText)
        practice_layout.addWidget(sentence_label)

        answer_box = QLineEdit()
        answer_box.setPlaceholderText("Enter the Chinese definition")
        answer_box.setObjectName("answerBox")
        answer_box.setInputMethodHints(Qt.InputMethodHint.ImhNone)
        answer_box.setMinimumWidth(700)
        answer_box.setFixedHeight(90)
        practice_layout.addWidget(answer_box, 0, Qt.AlignmentFlag.AlignHCenter)
        result_label = QLabel("")
        result_label.setObjectName("statusLabel")
        result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        result_label.setWordWrap(True)
        practice_layout.addWidget(result_label)
        report_hint = QLabel("")
        report_hint.setObjectName("reportHint")
        report_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        practice_layout.addWidget(report_hint)
        practice_layout.addStretch(1)

        action_row = QGridLayout()
        give_up_button = QPushButton("Give up")
        give_up_button.setObjectName("goButton")
        give_up_button.clicked.connect(self._give_up_current_word)
        action_row.addWidget(give_up_button, 0, 0, Qt.AlignmentFlag.AlignLeft)
        sentence_button = QPushButton("Sentence hint")
        sentence_button.setObjectName("goButton")
        sentence_button.clicked.connect(
            lambda: self._generate_sentence(word, explanation, sentence_button)
        )
        action_row.addWidget(sentence_button, 0, 1, Qt.AlignmentFlag.AlignCenter)
        ok_button = QPushButton("OK")
        ok_button.setObjectName("goButton")
        ok_button.clicked.connect(
            lambda: self._submit_answer(word, explanation, answer_box, ok_button)
        )
        answer_box.returnPressed.connect(ok_button.click)
        action_row.addWidget(ok_button, 0, 2, Qt.AlignmentFlag.AlignRight)
        action_row.setColumnStretch(0, 1)
        action_row.setColumnStretch(1, 1)
        action_row.setColumnStretch(2, 1)
        practice_layout.addLayout(action_row)

        self._set_central_widget(practice_widget)
        self.practice_widget = practice_widget
        self.practice_progress_label = progress_label
        self.practice_word_label = word_label
        self.practice_word_display = word_display
        self.sentence_label = sentence_label
        self.answer_box = answer_box
        self.ok_button = ok_button
        self.give_up_button = give_up_button
        self.sentence_button = sentence_button
        self.result_label = result_label
        self.report_hint = report_hint
        self.practice_layout = practice_layout
        answer_box.setFocus(Qt.FocusReason.OtherFocusReason)
        answer_box.activateWindow()

    def _submit_answer(
        self, word: str, explanation: str, answer_box: QLineEdit, ok_button: QPushButton
    ) -> None:
        answer = answer_box.text().strip()
        if not answer:
            self.result_label.setText("Enter an answer first")
            return
        self.give_up_button.hide()
        self.sentence_button.hide()
        self.practice_attempts += 1
        answer_box.setReadOnly(True)
        ok_button.setText("Judging...")
        ok_button.setEnabled(False)
        self.result_label.setText("Checking your answer...")

        thread = QThread(self)
        worker = AnswerJudgeWorker(word, explanation, answer)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._handle_judgment, Qt.ConnectionType.QueuedConnection)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        self.answer_judge_thread = thread
        self.answer_judge_worker = worker
        self.answer_judge_context = (answer_box, ok_button, thread, worker)
        thread.start()

    def _generate_sentence(
        self, word: str, explanation: str, sentence_button: QPushButton
    ) -> None:
        refresh = sentence_button.text() == "Refresh sentence"
        sentence_button.setText("Generating...")
        sentence_button.setEnabled(False)
        self.sentence_label.setText("Generating an example sentence...")

        thread = QThread(self)
        worker = SentenceWorker(word, explanation, refresh=refresh)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._handle_sentence, Qt.ConnectionType.QueuedConnection)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        self.sentence_thread = thread
        self.sentence_worker = worker
        self.sentence_context = (sentence_button, thread, worker)
        thread.start()

    @Slot(str, str)
    def _handle_sentence(self, sentence: str, error: str) -> None:
        context = getattr(self, "sentence_context", None)
        if context is None:
            return
        sentence_button, thread, worker = context
        if sentence:
            clean_sentence = re.sub(r"\[\[|\]\]", "", sentence)
            current_entry = self.practice_entries[self.practice_index]
            self.practice_sentence_hints[current_entry] = clean_sentence
            self.sentence_label.setText(self._highlight_tested_word(sentence, self.practice_word_display.text()))
            sentence_button.setText("Refresh sentence")
            sentence_button.setEnabled(True)
        else:
            self.sentence_label.setText(error)
            sentence_button.setText("Sentence hint")
            sentence_button.setEnabled(True)
        thread.finished.connect(lambda: self._release_sentence_references(worker))
        thread.quit()

    @staticmethod
    def _highlight_tested_word(sentence: str, word: str) -> str:
        marked_match = re.search(r"\[\[([^\]]+)\]\]", sentence)
        if marked_match:
            return (
                html.escape(sentence[:marked_match.start()])
                + '<span style="background-color: #fff176; color: #20211f;">'
                + html.escape(marked_match.group(1))
                + "</span>"
                + html.escape(sentence[marked_match.end():])
            )

        highlighted_parts: list[str] = []
        last_end = 0
        for match in re.finditer(re.escape(word), sentence, flags=re.IGNORECASE):
            highlighted_parts.append(html.escape(sentence[last_end:match.start()]))
            highlighted_parts.append(
                '<span style="background-color: #fff176; color: #20211f;">'
                f"{html.escape(match.group(0))}</span>"
            )
            last_end = match.end()
        highlighted_parts.append(html.escape(sentence[last_end:]))
        return "".join(highlighted_parts)

    @Slot(object, str)
    def _handle_judgment(self, correct: object, reason: str) -> None:
        context = getattr(self, "answer_judge_context", None)
        if context is None:
            return
        answer_box, ok_button, thread, worker = context
        self._show_judgment(
            correct, reason, answer_box, ok_button, thread, worker
        )
        thread.quit()

    def _show_judgment(
        self,
        correct: object,
        reason: str,
        answer_box: QLineEdit,
        ok_button: QPushButton,
        thread: QThread,
        worker: AnswerJudgeWorker,
    ) -> None:
        if correct is None:
            answer_box.setReadOnly(False)
            self.give_up_button.show()
            self.sentence_button.show()
            answer_box.setStyleSheet("border: 2px solid #c85132;")
            ok_button.setText("Try again")
            ok_button.setEnabled(True)
            self.result_label.setText(reason)
            return

        answer_box.setReadOnly(False)
        answer_box.setStyleSheet(
            "border: 2px solid #3a8f62;" if correct else "border: 2px solid #c85132;"
        )
        if correct:
            entry = self.practice_entries[self.practice_index]
            answer_box.setText(entry[1])
            answer_box.setReadOnly(True)
            self.give_up_button.hide()
            ok_button.setText("Next")
            ok_button.setEnabled(True)
            ok_button.setDefault(True)
            try:
                ok_button.clicked.disconnect()
            except RuntimeError:
                pass
            try:
                answer_box.returnPressed.disconnect(ok_button.click)
            except (RuntimeError, TypeError):
                pass
            ok_button.clicked.connect(
                lambda: QTimer.singleShot(0, self._show_next_practice_entry)
            )
            ok_button.setFocus(Qt.FocusReason.OtherFocusReason)
            self.result_label.setText(reason or "Correct")
        else:
            entry = self.practice_entries[self.practice_index]
            self.practice_error_counts[entry] += 1
            if self.practice_attempts < 2:
                answer_box.setReadOnly(False)
                answer_box.clear()
                answer_box.setStyleSheet("border: 2px solid #c85132;")
                answer_box.setFocus(Qt.FocusReason.OtherFocusReason)
                ok_button.setText("Try again")
                ok_button.setEnabled(True)
                self.result_label.setText("Incorrect. One attempt left.")
                return

            if entry not in self.practice_incorrect_entries:
                self.practice_incorrect_entries.append(entry)
            answer_box.setReadOnly(True)
            ok_button.setText("Next")
            ok_button.setEnabled(True)
            ok_button.setDefault(True)
            try:
                ok_button.clicked.disconnect()
            except RuntimeError:
                pass
            try:
                answer_box.returnPressed.disconnect(ok_button.click)
            except (RuntimeError, TypeError):
                pass
            ok_button.clicked.connect(
                lambda: QTimer.singleShot(0, self._show_next_practice_entry)
            )
            ok_button.setFocus(Qt.FocusReason.OtherFocusReason)
            self.result_label.setText(f"Incorrect twice. Correct answer: {entry[1]}")
            answer_box.setText(entry[1])
        thread.finished.connect(lambda: self._release_judge_references(worker))

    def _advance_after_answer(
        self,
        answer_box: QLineEdit,
        ok_button: QPushButton,
        thread: QThread,
        worker: AnswerJudgeWorker,
    ) -> None:
        pass

    def _show_next_practice_entry(self) -> None:
        self.practice_index += 1
        if self.practice_index >= len(self.practice_entries):
            if self.practice_incorrect_entries:
                self.practice_entries = self.practice_incorrect_entries
                self.practice_incorrect_entries = []
                self.practice_index = 0
                self.practice_reviewing = True
            else:
                self._finish_practice()
                return

        self.practice_attempts = 0
        self._show_practice_window(self.practice_entries[self.practice_index])

    def _give_up_current_word(self) -> None:
        entry = self.practice_entries[self.practice_index]
        if entry not in self.practice_given_up_entries:
            self.practice_given_up_entries.append(entry)
        self._show_next_practice_entry()

    def _finish_practice(self) -> None:
        self._show_statistics()

    def _show_statistics(self) -> None:
        export_entries = sorted(
            (
                entry[0],
                entry[1],
                "Given up"
                if entry in self.practice_given_up_entries
                else f"{self.practice_error_counts[entry]} incorrect",
                self.practice_sentence_hints.get(entry, ""),
            )
            for entry, count in self.practice_error_counts.items()
            if count > 0 or entry in self.practice_given_up_entries
        )
        statistics = "\n".join(
            f"{html.escape(entry[0])}: {count} incorrect"
            for entry, count in self.practice_error_counts.items()
            if entry not in self.practice_given_up_entries
        )
        given_up = "\n".join(
            f'<span style="color: #c85132;">{html.escape(entry[0])}</span> : unknown'
            for entry in self.practice_given_up_entries
        )
        sentence_hints = "\n".join(
            f"{html.escape(word)}: {html.escape(sentence)}"
            for word, sentence in sorted(
                (entry[0], sentence)
                for entry, sentence in self.practice_sentence_hints.items()
                if sentence
            )
        )
        report = f"Finished\n\n{statistics}"
        if given_up:
            report += f"\n\nGiven up\n{given_up}"
        if sentence_hints:
            report += f"\n\nSentence hints\n{sentence_hints}"
        self.practice_progress_label.hide()
        self.practice_word_label.hide()
        self.practice_word_display.hide()
        self.sentence_label.hide()
        self.answer_box.hide()
        self.give_up_button.hide()
        self.sentence_button.hide()
        self.ok_button.setText("Finish")
        self.ok_button.setEnabled(True)
        self.ok_button.setDefault(True)
        self.ok_button.setFocus(Qt.FocusReason.OtherFocusReason)
        try:
            self.ok_button.clicked.disconnect()
        except RuntimeError:
            pass
        self.ok_button.clicked.connect(self._return_to_start)
        self.result_label.setStyleSheet("font-size: 20px; font-weight: 600;")
        self.result_label.setTextFormat(Qt.TextFormat.RichText)
        self.result_label.setText(report.replace("\n", "<br>"))
        self.report_hint.setText("Press Enter or click Finish to start a new test")

        self.export_count_checkbox = QCheckBox("Include incorrect count in export")
        self.export_count_checkbox.setChecked(True)
        self.export_count_checkbox.setStyleSheet("font: 14px 'Avenir Next';")
        self.export_sentence_checkbox = QCheckBox("Include sentence hint in export")
        self.export_sentence_checkbox.setChecked(True)
        self.export_sentence_checkbox.setStyleSheet("font: 14px 'Avenir Next';")
        self.export_xlsx_button = QPushButton("Export .xlsx")
        self.export_xlsx_button.setObjectName("goButton")
        self.export_xlsx_button.clicked.connect(
            lambda: self._export_entries(export_entries, ".xlsx")
        )
        self.export_numbers_button = QPushButton("Export .numbers")
        self.export_numbers_button.setObjectName("goButton")
        self.export_numbers_button.clicked.connect(
            lambda: self._export_entries(export_entries, ".numbers")
        )
        export_layout = QHBoxLayout()
        export_layout.addWidget(self.export_count_checkbox)
        export_layout.addWidget(self.export_sentence_checkbox)
        export_layout.addStretch()
        export_layout.addWidget(self.export_xlsx_button)
        export_layout.addWidget(self.export_numbers_button)
        self.practice_layout.insertLayout(self.practice_layout.count() - 1, export_layout)

    def _export_entries(
        self, entries: list[tuple[str, str, str, str]], extension: str
    ) -> None:
        if not entries:
            self._show_standard_alert("Nothing to export", "No incorrect or given-up words.")
            return

        include_count = self.export_count_checkbox.isChecked()
        include_sentence = self.export_sentence_checkbox.isChecked()
        headers = ["Words", "Chinese definitions"]
        if include_count:
            headers.append("Result")
        if include_sentence:
            headers.append("Sentence hint")
        default_name = f"STM-review{extension}"
        output_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export review",
            str(Path.home() / default_name),
            f"{extension[1:].upper()} files (*{extension})",
        )
        if not output_path:
            return
        output_path = str(Path(output_path).with_suffix(extension))
        rows = [
            [
                word,
                definition,
                *([result] if include_count else []),
                *([sentence] if include_sentence else []),
            ]
            for word, definition, result, sentence in entries
        ]
        try:
            if extension == ".xlsx":
                workbook = Workbook()
                worksheet = workbook.active
                if worksheet is None:
                    raise RuntimeError("Could not create an Excel worksheet.")
                worksheet.title = "Review"
                worksheet.append(headers)
                for row in rows:
                    worksheet.append(row)
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.column_dimensions["A"].width = 28
                worksheet.column_dimensions["B"].width = 48
                if include_count:
                    worksheet.column_dimensions["C"].width = 18
                if include_sentence:
                    worksheet.column_dimensions["D" if include_count else "C"].width = 64
                workbook.save(output_path)
            else:
                document = NumbersDocument(
                    None,
                    sheet_name="Review",
                    table_name="Review",
                    num_header_rows=1,
                    num_header_cols=0,
                    num_rows=max(len(rows), 1),
                    num_cols=len(headers),
                )
                table = document.default_table
                for row_index, row in enumerate([headers, *rows]):
                    for column_index, value in enumerate(row):
                        table.write(row_index, column_index, value)
                document.save(output_path)
        except Exception as error:
            self._show_standard_alert("Export failed", str(error))
            return
        self._show_standard_alert("Export complete", f"Saved to {output_path}")

    def _return_to_start(self) -> None:
        if self.current_path is not None:
            try:
                self.current_path.unlink(missing_ok=True)
            except OSError:
                pass
        self.current_path = None
        self.words_table.setRowCount(0)
        self.words_label.setText("Words")
        self.file_label.setText("No file uploaded")
        self.storage_label.setText("Uploaded files are removed when you close STM.")
        self.status_label.setText("Waiting for a file to be uploaded...")
        self._set_go_file_state(False)
        self._set_central_widget(self.content_widget)
        self.menuBar().show()

    def _release_judge_references(self, worker: AnswerJudgeWorker) -> None:
        if getattr(self, "answer_judge_worker", None) is worker:
            self.answer_judge_worker = None
            self.answer_judge_thread = None
            self.answer_judge_context = None

    def _release_sentence_references(self, worker: SentenceWorker) -> None:
        if getattr(self, "sentence_worker", None) is worker:
            self.sentence_worker = None
            self.sentence_thread = None
            self.sentence_context = None

    def _set_go_file_state(self, has_file: bool) -> None:
        self.go_button.setProperty("hasFile", has_file)
        self.go_button.style().unpolish(self.go_button)
        self.go_button.style().polish(self.go_button)

    def _delete_word_row(self) -> None:
        button = self.sender()
        if button is None:
            return
        for row in range(self.words_table.rowCount()):
            if self.words_table.cellWidget(row, 2) is button:
                self.words_table.removeRow(row)
                self.words_label.setText(f"Words ({self.words_table.rowCount()})")
                return

    def _read_words(self, path: Path) -> list[tuple[str, str]]:
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                row_groups = [
                    list(worksheet.iter_rows(values_only=True))
                    for worksheet in workbook.worksheets
                ]
            finally:
                workbook.close()

        else:
            document = Document(str(path))
            row_groups = [
                list(table.rows(values_only=True))
                for sheet in document.sheets
                for table in sheet.tables
            ]

        rows = [
            [value for value in row]
            for group in row_groups
            for row in group
            if any(value not in (None, "") for value in row)
        ]
        if not rows:
            raise ValueError("The uploaded file does not contain any data.")

        return self._recognize_words_with_ai(rows)

    def _recognize_words_with_ai(self, rows: list[list[Any]]) -> list[tuple[str, str]]:
        api_key = (key or "").strip()
        if not api_key:
            raise ValueError("Please provide a DeepSeek API key first.")

        base_url = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
        model = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
        payload = {
            "model": model,
            "temperature": 0,
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "Extract English vocabulary words and their Chinese definitions from "
                        "the supplied spreadsheet rows. Do not rely on column names or their "
                        "language. Identify the two fields by their actual content. Ignore "
                        "titles, notes, numbering, empty rows, and unrelated columns. Return "
                        "JSON only in the form {\"entries\":[{\"word\":\"...\","
                        "\"definition\":\"...\"}]}. Keep the original text."
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps({"rows": rows}, ensure_ascii=False, default=str),
                },
            ],
        }
        request = urllib.request.Request(
            f"{base_url.rstrip('/')}/chat/completions",
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        try:
            result = self._request_json(request, base_url)
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, ssl.SSLError) as error:
            raise ValueError(f"Could not recognize words with DeepSeek: {error}") from error

        try:
            content = result["choices"][0]["message"]["content"].strip()
            if content.startswith("```"):
                content = content.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            recognized = json.loads(content)["entries"]
            entries = [
                (str(entry["word"]).strip(), str(entry["definition"]).strip())
                for entry in recognized
                if isinstance(entry, dict)
                and str(entry.get("word", "")).strip()
                and str(entry.get("definition", "")).strip()
            ]
        except (KeyError, IndexError, TypeError, json.JSONDecodeError) as error:
            raise ValueError("DeepSeek returned an invalid word recognition result.") from error

        if not entries:
            raise ValueError("DeepSeek could not find English words and Chinese definitions.")
        return entries

    def _request_json(self, request: urllib.request.Request, base_url: str) -> Any:
        try:
            with urllib.request.urlopen(
                request,
                timeout=30,
                context=AnswerJudgeWorker._build_ssl_context(base_url),
            ) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.URLError as error:
            if AnswerJudgeWorker._is_certificate_error(error):
                with urllib.request.urlopen(
                    request,
                    timeout=30,
                    context=ssl._create_unverified_context(),
                ) as response:
                    return json.loads(response.read().decode("utf-8"))
            raise

    def keyPressEvent(self, event: Any) -> None:
        if (
            event.key() == Qt.Key.Key_Escape
            and self.centralWidget() is self.content_widget
        ):
            self._return_to_key_entry()
            event.accept()
            return
        super().keyPressEvent(event)

    def closeEvent(self, event: Any) -> None:
        self.temporary_directory.cleanup()
        super().closeEvent(event)

    def _show_error(self, message: str) -> None:
        self.status_label.setText("Could not open file")
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Icon.NoIcon)
        box.setWindowTitle("Unable to open sheet")
        box.setText(message)
        box.setStandardButtons(QMessageBox.StandardButton.Ok)
        box.exec()


def main() -> None:
    application = QApplication(sys.argv)
    application.setApplicationName("STM")
    window = SpreadsheetWindow()
    window.show()
    sys.exit(application.exec())


if __name__ == "__main__":
    main()
